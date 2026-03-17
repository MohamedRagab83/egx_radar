"""
One-time script to generate a Paper Trading Excel tracker.
Run from project root:
    python -m egx_radar.tools.paper_trading_tracker

Generates: paper_trading_tracker.xlsx
"""

import pathlib
import sys

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Install openpyxl first: pip install openpyxl")
    sys.exit(1)


def create_tracker():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Paper Trades"

    # Header style
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center")

    headers = [
        "Date", "Symbol", "Sector", "Rank", "Zone",
        "Entry", "Stop", "Target", "R:R",
        "Price 1W", "Price 2W",
        "Result", "P&L %", "Notes"
    ]

    col_widths = [12, 8, 12, 8, 10, 8, 8, 8, 6, 10, 10, 10, 8, 20]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 25
    ws.freeze_panes = "A2"

    # Result dropdown (Data Validation)
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(
        type="list",
        formula1='"WIN,LOSS,HOLD,PENDING"',
        allow_blank=True,
        showDropDown=False,
    )
    ws.add_data_validation(dv)
    dv.sqref = "L2:L500"

    # Conditional formatting for Result column
    from openpyxl.styles.differential import DifferentialStyle
    from openpyxl.formatting.rule import Rule

    win_fill  = PatternFill("solid", fgColor="C6EFCE")
    loss_fill = PatternFill("solid", fgColor="FFC7CE")

    ws.conditional_formatting.add(
        "L2:L500",
        Rule(type="containsText", operator="containsText", text="WIN",
             dxf=DifferentialStyle(fill=win_fill))
    )
    ws.conditional_formatting.add(
        "L2:L500",
        Rule(type="containsText", operator="containsText", text="LOSS",
             dxf=DifferentialStyle(fill=loss_fill))
    )

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "EGX Radar — Paper Trading Summary"
    ws2["A1"].font = Font(bold=True, size=14)

    summary_labels = [
        ("A3", "Total Signals:"),
        ("A4", "WIN:"),
        ("A5", "LOSS:"),
        ("A6", "PENDING:"),
        ("A8", "Win Rate:"),
        ("A9", "Avg R:R (wins):"),
        ("A10", "Avg R:R (losses):"),
    ]
    for cell_ref, label in summary_labels:
        ws2[cell_ref] = label
        ws2[cell_ref].font = Font(bold=True)

    # Formulas referencing Paper Trades sheet
    ws2["B3"] = "=COUNTA('Paper Trades'!A2:A500)"
    ws2["B4"] = "=COUNTIF('Paper Trades'!L2:L500,\"WIN\")"
    ws2["B5"] = "=COUNTIF('Paper Trades'!L2:L500,\"LOSS\")"
    ws2["B6"] = "=COUNTIF('Paper Trades'!L2:L500,\"PENDING\")"
    ws2["B8"] = "=IF(B4+B5>0, B4/(B4+B5), 0)"
    ws2["B8"].number_format = "0.0%"

    output_path = pathlib.Path("paper_trading_tracker.xlsx")
    wb.save(output_path)
    print(f"✅ Paper Trading Tracker saved → {output_path.absolute()}")
    print("📋 Instructions:")
    print("   1. After each scan, copy PROBE signals from the daily CSV to this file")
    print("   2. After 1 week, fill in 'Price 1W' column")
    print("   3. Set Result to WIN/LOSS/HOLD")
    print("   4. Check Summary sheet for Win Rate")


if __name__ == "__main__":
    create_tracker()
